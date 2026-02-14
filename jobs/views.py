from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import JobDrive, ApplicationField, Application, ApplicationResponse
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from core.services.eligibility_service import check_student_eligibility
from django.contrib import messages
from students.models import StudentProfile
from core.models import GuidanceContent, StudentRoundStatus


# ==============================
# CREATE JOB
# ==============================
@login_required
def create_job(request):

    if request.user.role not in ['ADMIN', 'OFFICER']:
        return redirect('landing')

    if request.method == 'POST':

        job = JobDrive.objects.create(
            company_name=request.POST.get('company_name'),
            title=request.POST.get('title'),
            description=request.POST.get('description'),
            eligible_branches=request.POST.get('branches'),
            min_cgpa=request.POST.get('min_cgpa') or None,
            max_backlogs=request.POST.get('max_backlogs') or None,
            cgpa_required=request.POST.get('cgpa_required') == 'on',
            backlogs_required=request.POST.get('backlogs_required') == 'on',
            ctc=request.POST.get('salary_amount') or 0,
            card_color=request.POST.get('card_color'),
            image=request.FILES.get('company_logo'),
            application_deadline=request.POST.get('application_deadline'),
            created_by=request.user
        )

        # SAVE DYNAMIC FIELDS
        labels = request.POST.getlist('field_label[]')
        types = request.POST.getlist('field_type[]')

        for label, field_type in zip(labels, types):
            if label.strip():
                ApplicationField.objects.create(
                    job=job,
                    field_name=label,
                    field_type=field_type
                )

        return redirect('/dashboard/officer/')

    return render(request, 'jobs/create_job.html')


# ==============================
# JOB LIST
# ==============================
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import JobDrive, Application

@login_required
def job_list(request):

    jobs = JobDrive.objects.filter(status="OPEN")

    profile = getattr(request.user, "student_profile", None)

    # Get filter params
    selected_branch = request.GET.get("branch")
    eligibility_filter = request.GET.get("eligibility")

    # Branch filter
    if selected_branch and selected_branch != "all":
        jobs = jobs.filter(eligible_branches__icontains=selected_branch)

    jobs_with_data = []

    for job in jobs:
        eligibility = check_student_eligibility(profile, job)

        # Eligibility filter
        if eligibility_filter == "eligible" and not eligibility["eligible"]:
            continue
        if eligibility_filter == "not_eligible" and eligibility["eligible"]:
            continue
            
        is_applied = False
        if request.user.role == "STUDENT":
             is_applied = Application.objects.filter(job=job, student=request.user).exists()

        jobs_with_data.append({
            "job": job,
            "eligibility": eligibility,
            "is_applied": is_applied
        })

    # Get unique branches dynamically
    branch_list = set()
    for job in JobDrive.objects.all():
        if job.eligible_branches:
            for b in job.eligible_branches.split(","):
                branch_list.add(b.strip())

    context = {
        "jobs_with_data": jobs_with_data,
        "branch_list": sorted(branch_list),
        "selected_branch": selected_branch,
        "eligibility_filter": eligibility_filter,
    }

    return render(request, "jobs/job_list.html", context)



# ==============================
# JOB DETAIL
# ==============================
@login_required
def job_detail(request, job_id):

    job = get_object_or_404(JobDrive, id=job_id)
    already_applied = False
    
    eligibility = {"eligible": True, "reasons": []}
    profile = getattr(request.user, "student_profile", None)

    if request.user.role == 'STUDENT':
        
        eligibility = check_student_eligibility(profile, job)

        # Check for existing application
        is_applied = Application.objects.filter(job=job, student=request.user).exists()
        
        if is_applied:
            application = Application.objects.get(job=job, student=request.user)
            # mark viewed
            if not application.viewed:
                application.viewed = True
                application.save()
            already_applied = True 

    context = {
        'job': job,
        'fields': job.fields.all(),
        'already_applied': already_applied,
        'eligibility': eligibility,
        'profile': profile
    }
    if already_applied:
         context['application'] = Application.objects.get(job=job, student=request.user)

    return render(request, 'jobs/job_detail.html', context)


# ==============================
# APPLY JOB
# ==============================
@login_required
def apply_job(request, job_id):

    job = get_object_or_404(JobDrive, id=job_id)

    if job.application_deadline and timezone.now() > job.application_deadline:
        return render(request, 'jobs/job_closed.html', {'job': job})

    if request.user.role != 'STUDENT':
        return redirect('landing')

    # ELIGIBILITY CHECK
    profile = getattr(request.user, "student_profile", None)
    eligibility = check_student_eligibility(profile, job)

    if not eligibility["eligible"]:
        messages.error(request, "You are not eligible for this drive.")
        return redirect("job_detail", job_id=job.id)

    if request.method == 'POST':

        application, created = Application.objects.get_or_create(
            job=job,
            student=request.user
        )

        application.status = "APPLIED" # Updated status choice uppercase
        application.save()

        # Clear old responses
        ApplicationResponse.objects.filter(application=application).delete()

        for field in job.fields.all():
    
            if field.field_type in ["file", "multiple_file"]:
                uploaded_file = request.FILES.get(field.field_name)
                value = uploaded_file.name if uploaded_file else ""
            else:
                value = request.POST.get(field.field_name, "")
    
            ApplicationResponse.objects.create(
                application=application,
                field=field,
                value=value
            )

        return redirect('/dashboard/student/')

    return render(request, 'jobs/apply_job.html', {
        'job': job,
        'fields': job.fields.all()
    })


# ==============================
# UPDATE STATUS
# ==============================
@login_required
def update_status(request, application_id):

    if request.user.role not in ['ADMIN', 'OFFICER']:
        return redirect('landing')

    application = get_object_or_404(Application, id=application_id)

    if request.method == 'POST':
        status = request.POST.get('status')
        application.status = status
        application.save()

    return redirect('/dashboard/officer/')


# ==============================
# EXPORT EXCEL (Basic)
# ==============================
@login_required
def export_excel(request, job_id):

    job = get_object_or_404(JobDrive, id=job_id)
    applications = Application.objects.filter(job=job)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Applications"

    headers = ['Student Name', 'Roll Number', 'Branch', 'Status', 'Viewed']
    sheet.append(headers)

    for app in applications:
        sheet.append([
            app.student.username,
            app.student.roll_number, # This might fail if roll_number moved to profile, but existing User model had it. I should check User model. User model had it. I should prefer profile if I can, but standard User fields are safer for now.
            app.student.branch,
            app.status,
            "Yes" if app.viewed else "No"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={job.title}.xlsx'

    workbook.save(response)
    return response


# ==============================
# EXPORT DYNAMIC RESPONSES
# ==============================
@login_required
def export_student_responses(request, job_id):

    job = get_object_or_404(JobDrive, id=job_id)
    applications = Application.objects.filter(job=job)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dynamic Responses"

    dynamic_fields = job.fields.all()

    headers = [field.field_name for field in dynamic_fields]
    sheet.append(headers)

    for app in applications:
        row = []

        for field in dynamic_fields:
            response = ApplicationResponse.objects.filter(
                application=app,
                field=field
            ).first()

            row.append(response.value if response else "")

        sheet.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'{job.title}_dynamic_responses.xlsx'

    workbook.save(response)
    return response

# ==============================
# GUIDANCE SYSTEM
# ==============================
@login_required
def guidance_list(request):
    jobs = JobDrive.objects.all().order_by('-created_at')
    return render(request, 'jobs/guidance_list.html', {'jobs': jobs})

@login_required
def create_guidance(request, job_id):
    if request.user.role not in ["ADMIN", "OFFICER"]:
        return redirect("student_dashboard")

    job = get_object_or_404(JobDrive, id=job_id)

    if request.method == "POST":
        GuidanceContent.objects.create(
            job=job,
            round_name=request.POST.get("round_name"),
            title=request.POST.get("title"),
            resource_type=request.POST.get("resource_type"),
            file=request.FILES.get("file"),
            link=request.POST.get("link"),
            description=request.POST.get("description"),
            created_by=request.user,
        )
        return redirect("guidance_detail", job_id=job.id)

    return render(request, "jobs/create_guidance.html", {"job": job})

@login_required
def guidance_detail(request, job_id):
    job = get_object_or_404(JobDrive, id=job_id)
    
    guidance_contents = GuidanceContent.objects.filter(job=job).order_by('created_at')
    
    student_statuses = {}
    if request.user.role == "STUDENT":
        statuses = StudentRoundStatus.objects.filter(job=job, student=request.user)
        for s in statuses:
            student_statuses[s.round_name] = s.status

    # Group by round
    temp_grouped = {}
    for item in guidance_contents:
        if item.round_name not in temp_grouped:
            temp_grouped[item.round_name] = []
        temp_grouped[item.round_name].append(item)

    # Convert to list of objects for template
    final_guidance = []
    # Sort rounds if needed, currently alphabetical or insertion order?
    # Maybe use known rounds order if possible? 
    # For now, just sort by name or keep insertion order from guidance creation.
    # We can sort rounds by earliest created content in that round.
    
    # Sort rounds based on earliest created_at of content
    rounds_order = []
    seen = set()
    for item in guidance_contents:
        if item.round_name not in seen:
            rounds_order.append(item.round_name)
            seen.add(item.round_name)
            
    for round_name in rounds_order:
        contents = temp_grouped[round_name]
        status = student_statuses.get(round_name, "PENDING")
        
        final_guidance.append({
            "round_name": round_name,
            "contents": contents,
            "status": status
        })

    return render(request, "jobs/guidance_detail.html", {
        "job": job,
        "grouped_guidance": final_guidance, # Renamed variable in template usage
    })

@login_required
def update_round_status(request, job_id):
    if request.user.role != "STUDENT":
        return redirect("landing")

    if request.method == "POST":
        job = get_object_or_404(JobDrive, id=job_id)
        round_name = request.POST.get("round_name")
        status = request.POST.get("status")

        StudentRoundStatus.objects.update_or_create(
            student=request.user,
            job=job,
            round_name=round_name,
            defaults={"status": status}
        )
        
        return redirect("guidance_detail", job_id=job.id)
    
    return redirect("guidance_list")

@login_required
def export_job_guidance_data(request, job_id):

    if request.user.role not in ["ADMIN", "OFFICER"]:
        return redirect("student_dashboard")

    job = get_object_or_404(JobDrive, id=job_id)

    records = StudentRoundStatus.objects.filter(job=job)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Guidance Data"

    sheet.append(["Student", "Roll No", "Round", "Status"])

    for record in records:
        profile = getattr(record.student, 'student_profile', None)
        roll = profile.roll_number if profile else record.student.username
        sheet.append([
            record.student.username,
            roll,
            record.round_name,
            record.status,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={job.title}_guidance.xlsx"

    workbook.save(response)
    return response
